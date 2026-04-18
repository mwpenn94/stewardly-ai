"""
Build the AZ Lead Sourcing Segment-Channel Matrix as an xlsx workbook
matching the v5 family naming convention.

Creates: WealthBridge_Segment_Channel_Matrix_v5.xlsx
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path("/home/claude/wb/v6/v5_sample/WealthBridge_Segment_Channel_Matrix_v5.xlsx")

# Fit codes
FIT = {"high": "✅", "situational": "◑", "no": "❌"}

# Channel × segment data — pulled from AZ_LEAD_SOURCING_MASTER.md
SEGMENTS = ["Res Client", "Com Client", "Exp Pro", "New Assoc",
            "CPA/Atty", "Affiliate", "HR Dir", "Nonprof"]

CHANNELS = [
    # name, [fit codes per segment], cost-band, time-band, notes
    ("Existing client referrals",       ["high","high","sit","sit","high","high","sit","sit"], "$0-50",   "low",  "Highest LTV, lowest CAC; systematize the ask"),
    ("County parcel + outbound",        ["high","high","no","no","no","no","no","no"],         "$5-20",   "med",  "Current backbone for client segments"),
    ("AZ Corp Commission scrape",       ["no","high","no","no","sit","sit","no","no"],         "$3-10",   "low",  "AZ analog of NM SOS"),
    ("NM SOS / WA DOR scrape",          ["no","high","no","no","sit","sit","no","no"],         "$3-10",   "low",  "Free public business registries"),
    ("FINRA BrokerCheck",               ["no","no","high","no","no","sit","no","no"],          "$5-25",   "low",  "Free advisor data; throttled API"),
    ("AZ DOI licensee roster",          ["no","no","high","sit","no","high","no","no"],        "$5-25",   "low",  "Newly licensed = movability signal"),
    ("AZ State Bar directory",          ["no","no","no","no","high","no","no","no"],           "$10-30",  "med",  "Estate/Trust/Tax filter"),
    ("AZ Society of CPAs",              ["no","no","no","no","high","no","no","no"],           "$10-30",  "med",  "Annual symposium June"),
    ("AZ Dept of Real Estate",          ["sit","no","no","sit","sit","high","no","no"],        "$10-25",  "low",  "Realtors → Track A referral fees"),
    ("LinkedIn Sales Nav + Dripify",    ["sit","high","high","high","high","high","high","high"], "$20-60",  "med", "Already deployed in stack"),
    ("LinkedIn Jobs + Indeed (Workable)", ["no","no","sit","high","no","no","no","no"],         "$40-80",  "med",  "Inbound recruit funnel"),
    ("Davis-Monthan AFB TAP",           ["sit","no","sit","high","no","no","no","no"],         "$50-150", "high", "T1 Tucson; 360-600/yr separating"),
    ("Fort Huachuca TAP",               ["sit","no","sit","high","no","no","no","no"],         "$50-150", "high", "T2 Cochise; 75min from Tucson"),
    ("AZ Cattle Growers Assoc",         ["sit","high","no","no","no","no","no","sit"],         "$80-200", "med",  "Land-rich succession-heavy"),
    ("AZ Mining Assoc",                 ["sit","high","no","no","no","no","sit","no"],         "$80-200", "med",  "Sponsor + present"),
    ("AZ Tech Council",                 ["sit","high","no","sit","no","no","high","no"],       "$50-150", "med",  "Phoenix-tilted but valuable"),
    ("Hispanic Chamber AZ",             ["high","high","no","sit","no","sit","sit","high"],    "$30-100", "med",  "Strong Tucson/Nogales"),
    ("Greater Tucson Chamber",          ["high","high","sit","sit","high","sit","high","sit"], "$30-100", "med",  "Hub for T1 networking"),
    ("Sahuarita/Green Valley Chamber",  ["high","high","no","no","sit","no","sit","no"],       "$20-80",  "low",  "Retiree-tilted communities"),
    ("Sierra Vista Chamber",            ["high","high","no","no","sit","no","sit","no"],       "$20-80",  "low",  "T2 Cochise hub"),
    ("Nogales/Santa Cruz Chamber",      ["high","high","no","no","sit","no","no","no"],        "$20-80",  "low",  "T1 border community"),
    ("BNI chapters (4-6 in territory)", ["high","high","no","no","high","high","no","no"],     "$0-30",   "high", "Weekly attendance, peer trust"),
    ("Estate Planning Council Tucson",  ["sit","high","no","no","high","sit","no","no"],       "$10-80",  "med",  "Multi-disciplinary hub"),
    ("Rotary / Lions / Kiwanis",        ["high","high","no","no","sit","no","sit","high"],     "$0-60",   "high", "Sustained relational"),
    ("Faith communities",               ["high","high","no","sit","sit","no","no","high"],     "$0-30",   "med",  "Mike's church + COIs"),
    ("Workshops/seminars (T1)",         ["high","high","sit","sit","high","sit","high","sit"], "$150-400","high", "Closing engine"),
    ("CE-credit events for CPAs/Attys", ["no","no","no","no","high","no","no","no"],           "$200-400","med",  "You provide CE → relationship"),
    ("Meta lead ads",                   ["high","sit","no","sit","no","no","no","no"],         "$60-200", "low",  "Quality varies wildly"),
    ("Google Search (planning)",        ["high","high","no","no","no","no","no","no"],         "$80-300", "med",  "Inbound intent"),
    ("Google Search (recruiting)",      ["no","no","high","high","no","no","no","no"],         "$80-300", "med",  "Recruit-side"),
    ("SEC EDGAR Form 4",                ["high","high","no","no","no","no","no","no"],         "$5-15",   "low",  "Highest-propensity wealth trigger"),
    ("FAA aircraft registry",           ["high","high","no","no","no","no","no","no"],         "$5-25",   "low",  "Wealth signal enrichment"),
    ("County recorder (refi/lien)",     ["high","sit","no","no","no","no","no","no"],          "$5-15",   "med",  "Liquidity event timing"),
    ("Probate court (inheritance)",     ["high","sit","no","no","sit","no","no","no"],         "$10-40",  "med",  "Sensitive — relational"),
    ("MLS / Zillow (recent ≥$1M)",      ["high","no","no","no","no","sit","no","no"],          "$10-30",  "low",  "Buy + sell side"),
    ("NPI Registry (medical pros)",     ["high","high","no","no","sit","no","high","no"],      "$10-60",  "low",  "High LTV professionals"),
    ("GuideStar (nonprofit data)",      ["no","no","no","no","no","no","no","high"],           "$5-20",   "low",  "Nonprofit segment foundation"),
    ("AZ employer registry",            ["no","high","no","no","no","no","high","no"],         "$10-80",  "med",  "Workers' comp + UI source"),
]

# Styles
HDR = Font(bold=True, color="FFFFFF", size=11)
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

wb = Workbook()

# === Sheet 1: Cover ===
cover = wb.active
cover.title = "Cover"
cover["A1"] = "WealthBridge Segment-Channel Matrix v5"
cover["A1"].font = Font(bold=True, size=18, color="1F4E78")
cover["A3"] = "Companion to: AZ_LEAD_SOURCING_MASTER.md, GHL_AUTOMATION_SPEC_v2.md"
cover["A4"] = "Generated:  2026-04-15"
cover["A6"] = "Reading Guide:"
cover["A7"] = "  ✅ = high fit  |  ◑ = situational  |  ❌ = wrong tool"
cover["A8"] = "  Cost = per qualified conversation, not per lead"
cover["A9"] = "  Time = effort to set up + maintain (low/med/high)"
cover["A11"] = "Sheets:"
cover["A12"] = "  1. Cover (this sheet)"
cover["A13"] = "  2. Channel Matrix — full 38 channels × 8 segments"
cover["A14"] = "  3. By Segment — top 5 channels per segment, ranked"
cover["A15"] = "  4. Triggers — wealth-trigger event sources only"
cover["A16"] = "  5. Compliance Notes — segment-specific gates"
for col in ["A"]:
    cover.column_dimensions[col].width = 90

# === Sheet 2: Channel Matrix ===
ws = wb.create_sheet("Channel Matrix")
headers = ["Channel"] + SEGMENTS + ["Cost band", "Time", "Notes"]
for i, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=i, value=h)
    c.font = HDR; c.fill = HDR_FILL; c.alignment = ALIGN_CENTER; c.border = BORDER
ws.row_dimensions[1].height = 30
for r, (name, fits, cost, time, notes) in enumerate(CHANNELS, 2):
    ws.cell(row=r, column=1, value=name).alignment = ALIGN_LEFT
    for i, fit in enumerate(fits):
        cell = ws.cell(row=r, column=2 + i, value=FIT["high" if fit == "high" else "situational" if fit == "sit" else "no"])
        cell.alignment = ALIGN_CENTER
        if fit == "high":
            cell.fill = PatternFill("solid", fgColor="C6EFCE")
        elif fit == "sit":
            cell.fill = PatternFill("solid", fgColor="FFEB9C")
    ws.cell(row=r, column=10, value=cost).alignment = ALIGN_CENTER
    ws.cell(row=r, column=11, value=time).alignment = ALIGN_CENTER
    ws.cell(row=r, column=12, value=notes).alignment = ALIGN_LEFT
    for col in range(1, 13):
        ws.cell(row=r, column=col).border = BORDER
ws.column_dimensions["A"].width = 38
for col_idx in range(2, 10): ws.column_dimensions[get_column_letter(col_idx)].width = 12
ws.column_dimensions["J"].width = 12
ws.column_dimensions["K"].width = 8
ws.column_dimensions["L"].width = 50
ws.freeze_panes = "B2"
ws.auto_filter.ref = ws.dimensions

# === Sheet 3: By Segment ===
ws3 = wb.create_sheet("By Segment")
ws3["A1"] = "Top 5 Channels per Segment (priority order)"
ws3["A1"].font = Font(bold=True, size=14)
top_by_seg = {
    "Residential Client": ["1. Existing client referrals", "2. County parcel + outbound", "3. SEC Form 4 + probate + refi triggers", "4. Local chambers + faith + BNI", "5. Workshops"],
    "Commercial Client":  ["1. AZ Corp Commission + NAICS outbound", "2. Existing client + COI referrals", "3. Industry assocs (Cattle, Mining, Tech)", "4. LinkedIn Sales Nav (titled)", "5. Workshops + CE events"],
    "Experienced Pro":    ["1. FINRA BrokerCheck + LinkedIn cross-ref", "2. AZ DOI licensee roster", "3. LinkedIn outbound via Dripify", "4. Estate Planning Council / NAIFA", "5. Industry-specific recruiting"],
    "New Associate":      ["1. LinkedIn Jobs + Indeed via Workable", "2. Davis-Monthan + Fort Huachuca TAP", "3. UA Eller / ASU career services", "4. LinkedIn outbound (teachers/coaches/officers)", "5. DAV / VFW / AmVets"],
    "CPA / Attorney":     ["1. AZ State Bar + ASCPA directories", "2. Estate Planning Council Tucson", "3. CE-credit workshops", "4. LinkedIn + Dripify (per orchestration)", "5. BNI / Rotary / faith"],
    "Affiliate":          ["1. AZ DOI licensee roster + filtering", "2. AZ Dept of Real Estate licensees", "3. P&C agency owner outreach", "4. LinkedIn outbound (Insurance Agent + Indep)", "5. Existing affiliate referrals"],
    "HR Director":        ["1. AZ employer registry (50+ employees)", "2. LinkedIn Sales Nav (HR/Benefits/People Ops)", "3. Open-enrollment workshops (Aug-Nov)", "4. Industry chambers (Tucson, Tech, Mining)", "5. Healthcare HR networks (NPI cross-ref)"],
    "Nonprofit Leader":   ["1. GuideStar / Candid (AZ + revenue tier)", "2. AZ Community Foundation grantees", "3. Faith + Rotary + Lions clubs", "4. UA Eller Center Nonprofit Leadership", "5. Hispanic Chamber + cultural orgs"],
}
row = 3
for seg, items in top_by_seg.items():
    c = ws3.cell(row=row, column=1, value=seg); c.font = Font(bold=True, size=12); c.fill = SECTION_FILL
    row += 1
    for item in items:
        ws3.cell(row=row, column=1, value=item)
        row += 1
    row += 1
ws3.column_dimensions["A"].width = 60

# === Sheet 4: Triggers ===
ws4 = wb.create_sheet("Triggers")
ws4["A1"] = "Wealth-Trigger Event Sources — Highest Propensity"
ws4["A1"].font = Font(bold=True, size=14)
trig_headers = ["Source", "Trigger Type", "Volume (AZ)", "Lead-to-Conv", "Notes"]
for i, h in enumerate(trig_headers, 1):
    c = ws4.cell(row=3, column=i, value=h); c.font = HDR; c.fill = HDR_FILL; c.alignment = ALIGN_CENTER
TRIGGERS = [
    ("SEC Form 4 (insider sales ≥$1M)", "Liquidity event", "5-15/mo",   "8-15%",  "Public-co execs, very HNW"),
    ("County recorder (refi ≥$500K)",   "Cash-out event",  "30-80/mo",  "3-8%",   "Implies equity availability"),
    ("Probate filings (estates ≥$1M)",  "Inheritance",     "20-40/mo",  "5-12%",  "Sensitive — relational approach"),
    ("Divorce filings (high-asset)",    "Restructure",     "15-30/mo",  "2-5%",   "Sensitive timing"),
    ("Business sale filings",           "Exit liquidity",  "10-25/mo",  "10-20%", "Highest-LTV residential"),
    ("MLS sales ≥$1M",                  "RE liquidity",    "40-100/mo", "2-6%",   "Buy + sell side"),
    ("FAA aircraft registration",       "Wealth signal",   "5-15/mo",   "n/a",    "Enrichment join, not lead source"),
]
for r, row_data in enumerate(TRIGGERS, 4):
    for c, v in enumerate(row_data, 1):
        cell = ws4.cell(row=r, column=c, value=v)
        cell.alignment = ALIGN_LEFT if c in (1, 5) else ALIGN_CENTER
        cell.border = BORDER
ws4.column_dimensions["A"].width = 35
ws4.column_dimensions["B"].width = 18
ws4.column_dimensions["C"].width = 14
ws4.column_dimensions["D"].width = 14
ws4.column_dimensions["E"].width = 50

# === Sheet 5: Compliance Notes ===
ws5 = wb.create_sheet("Compliance Notes")
ws5["A1"] = "Segment-Specific Compliance Gates"
ws5["A1"].font = Font(bold=True, size=14)
comp_headers = ["Segment", "Gate"]
for i, h in enumerate(comp_headers, 1):
    c = ws5.cell(row=3, column=i, value=h); c.font = HDR; c.fill = HDR_FILL; c.alignment = ALIGN_CENTER
COMP = [
    ("experienced_pro",     "Transition-capital language is [PENDING TC APPROVAL] — keep in Draft until ESI signs off"),
    ("cpa_attorney_partner","Revenue-sharing language flagged [PENDING TC APPROVAL] until compliance reviews per state"),
    ("affiliate",           "Track B/C/D agreements require state-by-state confirmation that referral fees are permissible"),
    ("hr_director",         "Group benefits require AZ Group L&H endorsement on producer license; verify before workshop"),
    ("nonprofit_leader",    "Charitable planning content (CGAs, CRTs) requires advanced markets review"),
    ("ALL recruit + partner","Use only Track A safe content from linkedin_posts_enhanced.md for unattended automation"),
    ("ALL outbound",        "DNC scrub required before any call campaign; SMS requires 10DLC registration + STOP keyword"),
    ("ALL marketing",       "ESI compliance review 5-10 business days lead time per piece"),
]
for r, row_data in enumerate(COMP, 4):
    for c, v in enumerate(row_data, 1):
        cell = ws5.cell(row=r, column=c, value=v)
        cell.alignment = ALIGN_LEFT
        cell.border = BORDER
ws5.column_dimensions["A"].width = 24
ws5.column_dimensions["B"].width = 100

wb.save(OUT)
print(f"✓ wrote {OUT}")
print(f"  Sheets: {wb.sheetnames}")
print(f"  Channel matrix: {len(CHANNELS)} channels × {len(SEGMENTS)} segments = {len(CHANNELS)*len(SEGMENTS)} cells")

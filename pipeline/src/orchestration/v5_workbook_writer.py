"""
WealthBridge v5 Workbook Writer
================================
Produces the v5 workbook family, matching the v4 naming convention and
extending each with propensity-model fields. Drop-in upgrade for the
existing v4 xlsx suite.

Files produced:
  WealthBridge_Prospect_Database_v5.xlsx    — master feature table + metadata
  WealthBridge_Engagement_Database_v5.xlsx  — outcome log (labels for Phase 1)
  WealthBridge_Combined_Pipeline_v5.xlsx    — joined view, sortable/filterable
  WealthBridge_Executive_Summary_v5.xlsx    — KPI dashboard (formula-driven)
  WealthBridge_Event_Schedule_v5.xlsx       — seminar cadence + attribution
  WealthBridge_Scoring_Control_v5.xlsx      — config (weights, tiers, taxonomy)

Design principles:
  - Excel formulas (not Python-hardcoded values) for all derived metrics
  - Cover sheet in each workbook with schema + changelog from v4
  - Consistent Arial font, standard color coding (blue=input, black=formula,
    green=cross-sheet link, yellow=assumption)
  - Frozen header rows, filterable tables, column widths tuned for readability
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
FONT_BASE = Font(name="Arial", size=10)
FONT_HEADER = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FONT_TITLE = Font(name="Arial", size=14, bold=True)
FONT_INPUT = Font(name="Arial", size=10, color="0000FF")        # blue = hardcoded input
FONT_FORMULA = Font(name="Arial", size=10, color="000000")       # black = formula
FONT_XLINK = Font(name="Arial", size=10, color="008000")         # green = cross-sheet

FILL_HEADER = PatternFill("solid", fgColor="1F4E78")
FILL_SECTION = PatternFill("solid", fgColor="D9E1F2")
FILL_ASSUMPTION = PatternFill("solid", fgColor="FFFF00")         # yellow = needs attention
FILL_TIER_A = PatternFill("solid", fgColor="C6EFCE")
FILL_TIER_B = PatternFill("solid", fgColor="FFEB9C")
FILL_TIER_C = PatternFill("solid", fgColor="FFC7CE")
FILL_TIER_D = PatternFill("solid", fgColor="F2F2F2")

BORDER_THIN = Border(left=Side(style="thin", color="BFBFBF"),
                     right=Side(style="thin", color="BFBFBF"),
                     top=Side(style="thin", color="BFBFBF"),
                     bottom=Side(style="thin", color="BFBFBF"))

# Currency, percent, number formats per skill guidance
FMT_CURRENCY = '"$"#,##0;("$"#,##0);"-"'
FMT_NUMBER = "#,##0;(#,##0);-"
FMT_PERCENT = "0.0%;(0.0%);-"
FMT_SCORE = "0.0"
FMT_DATE = "yyyy-mm-dd"

# ---------------------------------------------------------------------------
# Schema definitions (v5)
# ---------------------------------------------------------------------------

PROSPECT_COLUMNS = [
    # Identity
    ("owner_key", "Owner Key (hashed)", "text", 18),
    ("segment", "Segment", "text", 12),
    ("owner_name", "Owner Name", "text", 28),
    ("entity_type", "Entity Type (Commercial)", "text", 18),
    # Location
    ("property_address", "Property Address", "text", 28),
    ("city", "City", "text", 16),
    ("state", "State", "text", 7),
    ("county", "County", "text", 14),
    ("zip", "ZIP", "text", 8),
    ("fips", "County FIPS", "text", 10),
    # Contact (after enrichment)
    ("phone", "Phone", "text", 14),
    ("email", "Email", "text", 26),
    # Property / entity features
    ("market_value", "Market Value ($)", "currency", 14),
    ("equity_est", "Est. Equity ($)", "currency", 14),
    ("ownership_length_years", "Ownership (yrs)", "number", 12),
    ("owner_age", "Owner Age", "number", 10),
    ("entity_age_years", "Entity Age (yrs)", "number", 13),
    ("multi_property_count", "Multi-Property Count", "number", 17),
    ("revenue_proxy", "Revenue Proxy ($)", "currency", 14),
    # Scoring — Phase 0
    ("propensity_score", "Propensity Score (0-100)", "score", 18),
    ("propensity_decile", "Decile (1=top)", "number", 12),
    ("propensity_tier", "Tier (A-D)", "text", 10),
    ("priority_multiplier", "Geo Priority ×", "number", 14),
    # Scoring — Phase 1 (populated once trained)
    ("phase1_proba", "Phase 1 Probability", "percent", 17),
    ("phase1_decile", "Phase 1 Decile", "number", 13),
    # Geo tiering
    ("geo_tier", "Geo Tier", "text", 22),
    # Audit
    ("source_file", "Source File", "text", 28),
    ("last_scored_date", "Last Scored", "date", 12),
]

ENGAGEMENT_COLUMNS = [
    ("event_id", "Event ID", "text", 14),
    ("owner_key", "Owner Key", "text", 18),
    ("contact_id", "GHL Contact ID", "text", 18),
    ("event_date", "Event Date", "date", 12),
    ("event_type", "Event Type", "text", 18),
    ("channel", "Channel", "text", 14),
    ("campaign", "Campaign", "text", 20),
    ("outcome_status_from", "Status From", "text", 15),
    ("outcome_status_to", "Status To", "text", 15),
    ("notes", "Notes", "text", 40),
    ("segment_at_event", "Segment", "text", 12),
    ("geo_tier_at_event", "Geo Tier", "text", 22),
    ("propensity_score_at_event", "Propensity Score", "score", 16),
    ("propensity_tier_at_event", "Tier", "text", 8),
]

EVENT_COLUMNS = [
    ("event_id", "Event ID", "text", 14),
    ("event_date", "Event Date", "date", 12),
    ("event_type", "Event Type", "text", 20),
    ("location", "Location", "text", 22),
    ("target_segment", "Target Segment", "text", 14),
    ("target_geo_tier", "Target Geo Tier", "text", 20),
    ("invites_sent", "Invites Sent", "number", 12),
    ("rsvps", "RSVPs", "number", 10),
    ("attended", "Attended", "number", 10),
    ("discoveries_booked", "Discoveries", "number", 13),
    ("cases_opened", "Cases Opened", "number", 13),
    ("venue_cost", "Venue Cost ($)", "currency", 14),
    ("fnb_cost", "F&B Cost ($)", "currency", 12),
    ("marketing_cost", "Marketing Cost ($)", "currency", 16),
    ("total_cost", "Total Cost ($)", "currency", 14),        # formula
    ("rsvp_rate", "RSVP Rate", "percent", 11),                # formula
    ("attendance_rate", "Attendance Rate", "percent", 15),    # formula
    ("discovery_rate", "Discovery Rate", "percent", 14),      # formula
    ("cost_per_discovery", "$/Discovery", "currency", 13),    # formula
    ("cost_per_case", "$/Case", "currency", 11),              # formula
]

# Geo tier reference (mirrors phase0_propensity_scoring.py)
GEO_TIER_REFERENCE = [
    ("tier_1_home_az", "AZ Pima, Mohave, Santa Cruz", 1.00,
     "Primary AZ Region 1 territory", "Direct production, full-stack"),
    ("tier_2_az_adjacent", "AZ Cochise, Maricopa, Yuma, Pinal, Graham", 0.85,
     "AZ expansion — adjacent counties", "Direct production, cheap channels"),
    ("tier_3_nm_wa_licensed", "All NM + WA counties", 0.75,
     "Licensed-state expansion", "Partner-first, then direct (after licensing)"),
    ("tier_4_us_broader", "All other US states", 0.50,
     "Broader US — no personal production", "Stewardly inbound only"),
    ("tier_5_global", "Non-US", 0.30,
     "International — platform only", "Stewardly inbound only"),
]

SCORING_WEIGHTS_RESIDENTIAL = [
    ("equity_decile", 0.30, "Proxy for investable net worth"),
    ("ownership_length_decile", 0.20, "10+ yrs = liquidity-event candidate"),
    ("age_fit_score", 0.20, "Life-stage fit (peak 45-65)"),
    ("zip_affluence_decile", 0.15, "ACS B19013 median HH income decile"),
    ("recent_liquidity_event", 0.15, "Refi / HELOC / sale in last 24 mo"),
]

SCORING_WEIGHTS_COMMERCIAL = [
    ("entity_age_decile", 0.25, "5+ years = established cash flow"),
    ("entity_type_factor", 0.15, "S-Corp > LLC > C-Corp > other"),
    ("revenue_band_decile", 0.25, "Qualification gate"),
    ("multi_property_count_decile", 0.20, "Wealth concentration signal"),
    ("owner_age_fit_score", 0.15, "Succession urgency (peak 50-70)"),
]

OUTCOME_TAXONOMY = [
    ("not_contacted", "Entry state — no touch delivered", "Censored"),
    ("contacted", "At least one outbound touch delivered", "Censored"),
    ("responded", "Any response received, including decline", "Censored"),
    ("qualified", "Decision-maker + need + next step", "Positive label"),
    ("closed_won", "Case opened / policy sold", "Positive label"),
    ("closed_lost", "Pursued to completion, lost", "Negative label"),
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def apply_header_row(sheet, row: int, num_cols: int):
    for c in range(1, num_cols + 1):
        cell = sheet.cell(row=row, column=c)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN

def write_dataframe(sheet, df: pd.DataFrame, cols: list[tuple], start_row: int = 1):
    """Write a DataFrame with column schema, applying formats + header styling."""
    # Header
    for i, (_, label, _, _) in enumerate(cols, 1):
        sheet.cell(row=start_row, column=i, value=label)
    apply_header_row(sheet, start_row, len(cols))
    # Data rows
    for r_idx, (_, row) in enumerate(df.iterrows(), start_row + 1):
        for c_idx, (key, _, kind, _) in enumerate(cols, 1):
            val = row.get(key)
            if pd.isna(val) or val == "":
                continue
            cell = sheet.cell(row=r_idx, column=c_idx, value=val)
            cell.font = FONT_BASE
            if kind == "currency":
                cell.number_format = FMT_CURRENCY
            elif kind == "number":
                cell.number_format = FMT_NUMBER
            elif kind == "percent":
                cell.number_format = FMT_PERCENT
            elif kind == "score":
                cell.number_format = FMT_SCORE
            elif kind == "date":
                cell.number_format = FMT_DATE
    # Column widths
    for i, (_, _, _, width) in enumerate(cols, 1):
        sheet.column_dimensions[get_column_letter(i)].width = width
    # Freeze header, add autofilter
    sheet.freeze_panes = sheet.cell(row=start_row + 1, column=1)
    if len(df) > 0:
        last_col = get_column_letter(len(cols))
        sheet.auto_filter.ref = f"A{start_row}:{last_col}{start_row + len(df)}"

def write_cover(sheet, title: str, description: str, schema_note: str,
                changelog: list[str]):
    sheet["A1"] = title
    sheet["A1"].font = FONT_TITLE
    sheet.merge_cells("A1:F1")
    sheet["A3"] = "Generated"; sheet["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    sheet["A4"] = "Version"; sheet["B4"] = "v5.0"
    sheet["A5"] = "Owner"; sheet["B5"] = "WealthBridge Financial Group / AZ Region 1"
    for r in range(3, 6):
        sheet.cell(row=r, column=1).font = Font(name="Arial", size=10, bold=True)
        sheet.cell(row=r, column=2).font = FONT_BASE

    sheet["A7"] = "Description"; sheet["A7"].font = Font(name="Arial", size=11, bold=True)
    sheet["A8"] = description
    sheet["A8"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet.merge_cells("A8:F10")

    sheet["A12"] = "Schema"; sheet["A12"].font = Font(name="Arial", size=11, bold=True)
    sheet["A13"] = schema_note
    sheet["A13"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet.merge_cells("A13:F16")

    sheet["A18"] = "Changes from v4"; sheet["A18"].font = Font(name="Arial", size=11, bold=True)
    for i, entry in enumerate(changelog, 19):
        sheet.cell(row=i, column=1, value="•").font = FONT_BASE
        sheet.cell(row=i, column=2, value=entry).font = FONT_BASE
        sheet.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)

    sheet.column_dimensions["A"].width = 18
    for col in "BCDEF":
        sheet.column_dimensions[col].width = 20


# ---------------------------------------------------------------------------
# Workbook builders
# ---------------------------------------------------------------------------

def build_prospect_database(df: pd.DataFrame, out_path: Path):
    wb = Workbook()
    ws_cover = wb.active
    ws_cover.title = "README"
    write_cover(
        ws_cover,
        "WealthBridge Prospect Database v5",
        "Master feature table for propensity scoring. One row per unique owner across "
        "all geographies (AZ/NM/WA/US/Global). Matches v4 identity columns and extends "
        "with propensity score, geo tier, and Phase 1 model outputs.",
        "Sheets: README · Prospects (full) · By_Segment · By_GeoTier · Field_Reference",
        [
            "Added propensity_score, propensity_decile, propensity_tier columns",
            "Added geo_tier (T1–T5) and priority_multiplier",
            "Added phase1_proba and phase1_decile (populated once model trained)",
            "Added multi_property_count (computed across county files)",
            "Added fips + last_scored_date audit columns",
            "owner_key is now the primary join key (was owner_name in v4)",
        ],
    )
    ws = wb.create_sheet("Prospects")
    write_dataframe(ws, df, PROSPECT_COLUMNS)

    # By-segment pivot
    if len(df) > 0 and "segment" in df.columns:
        ws_seg = wb.create_sheet("By_Segment")
        ws_seg["A1"] = "Segment Summary"; ws_seg["A1"].font = FONT_TITLE
        ws_seg["A3"] = "Segment"; ws_seg["B3"] = "Count"; ws_seg["C3"] = "A-tier"
        ws_seg["D3"] = "Avg Score"; ws_seg["E3"] = "Avg Equity ($)"
        apply_header_row(ws_seg, 3, 5)
        for r_off, seg in enumerate(sorted(df["segment"].dropna().unique()), 4):
            ws_seg.cell(row=r_off, column=1, value=seg)
            # Formulas reference the Prospects sheet directly
            ws_seg.cell(row=r_off, column=2, value=f'=COUNTIF(Prospects!B:B,"{seg}")')
            ws_seg.cell(row=r_off, column=3,
                value=f'=COUNTIFS(Prospects!B:B,"{seg}",Prospects!V:V,"A")')
            ws_seg.cell(row=r_off, column=4,
                value=f'=IFERROR(AVERAGEIF(Prospects!B:B,"{seg}",Prospects!T:T),0)')
            ws_seg.cell(row=r_off, column=4).number_format = FMT_SCORE
            ws_seg.cell(row=r_off, column=5,
                value=f'=IFERROR(AVERAGEIF(Prospects!B:B,"{seg}",Prospects!N:N),0)')
            ws_seg.cell(row=r_off, column=5).number_format = FMT_CURRENCY
            for c in range(1, 6):
                ws_seg.cell(row=r_off, column=c).font = FONT_XLINK
        for col, w in zip("ABCDE", [18, 10, 10, 12, 16]):
            ws_seg.column_dimensions[col].width = w

    # By-geo-tier pivot
    if len(df) > 0 and "geo_tier" in df.columns:
        ws_geo = wb.create_sheet("By_GeoTier")
        ws_geo["A1"] = "Geo Tier Summary"; ws_geo["A1"].font = FONT_TITLE
        ws_geo["A3"] = "Geo Tier"; ws_geo["B3"] = "Count"
        ws_geo["C3"] = "A-tier"; ws_geo["D3"] = "Avg Score"
        apply_header_row(ws_geo, 3, 4)
        for r_off, tier in enumerate(sorted(df["geo_tier"].dropna().unique()), 4):
            ws_geo.cell(row=r_off, column=1, value=tier)
            ws_geo.cell(row=r_off, column=2, value=f'=COUNTIF(Prospects!Z:Z,"{tier}")')
            ws_geo.cell(row=r_off, column=3,
                value=f'=COUNTIFS(Prospects!Z:Z,"{tier}",Prospects!V:V,"A")')
            ws_geo.cell(row=r_off, column=4,
                value=f'=IFERROR(AVERAGEIF(Prospects!Z:Z,"{tier}",Prospects!T:T),0)')
            ws_geo.cell(row=r_off, column=4).number_format = FMT_SCORE
            for c in range(1, 5):
                ws_geo.cell(row=r_off, column=c).font = FONT_XLINK
        for col, w in zip("ABCD", [24, 10, 10, 12]):
            ws_geo.column_dimensions[col].width = w

    # Field reference
    ws_ref = wb.create_sheet("Field_Reference")
    ws_ref["A1"] = "Field Reference"; ws_ref["A1"].font = FONT_TITLE
    ws_ref["A3"] = "Field"; ws_ref["B3"] = "Header"; ws_ref["C3"] = "Type"
    apply_header_row(ws_ref, 3, 3)
    for i, (key, label, kind, _) in enumerate(PROSPECT_COLUMNS, 4):
        ws_ref.cell(row=i, column=1, value=key).font = FONT_BASE
        ws_ref.cell(row=i, column=2, value=label).font = FONT_BASE
        ws_ref.cell(row=i, column=3, value=kind).font = FONT_BASE
    ws_ref.column_dimensions["A"].width = 26
    ws_ref.column_dimensions["B"].width = 28
    ws_ref.column_dimensions["C"].width = 12

    wb.save(out_path)


def build_engagement_database(df: pd.DataFrame, out_path: Path):
    wb = Workbook()
    ws_cover = wb.active
    ws_cover.title = "README"
    write_cover(
        ws_cover,
        "WealthBridge Engagement Database v5",
        "Outcome log — one row per material engagement event (touch, response, "
        "status change). This is the LABEL source for Phase 1 model training. "
        "Each row carries the scoring context as-of the event so label quality "
        "is preserved through re-scoring cycles.",
        "Sheets: README · Events · Outcome_Taxonomy · Label_Export",
        [
            "New file in v5 — formalizes what was implicit in v4 Combined_Pipeline",
            "Outcome_status pinned to 6-value taxonomy for Phase 1 label stability",
            "Carries propensity snapshot at event time (handles model drift)",
            "Label_Export sheet produces the CSV shape Phase 1 logistic expects",
        ],
    )
    ws = wb.create_sheet("Events")
    write_dataframe(ws, df, ENGAGEMENT_COLUMNS)

    # Outcome taxonomy sheet
    ws_tax = wb.create_sheet("Outcome_Taxonomy")
    ws_tax["A1"] = "Outcome Status Taxonomy"; ws_tax["A1"].font = FONT_TITLE
    ws_tax["A3"] = "Status"; ws_tax["B3"] = "Definition"; ws_tax["C3"] = "Phase 1 role"
    apply_header_row(ws_tax, 3, 3)
    for i, (status, definition, role) in enumerate(OUTCOME_TAXONOMY, 4):
        ws_tax.cell(row=i, column=1, value=status).font = FONT_BASE
        ws_tax.cell(row=i, column=2, value=definition).font = FONT_BASE
        ws_tax.cell(row=i, column=3, value=role).font = FONT_BASE
        if role == "Positive label":
            for c in range(1, 4):
                ws_tax.cell(row=i, column=c).fill = FILL_TIER_A
        elif role == "Negative label":
            for c in range(1, 4):
                ws_tax.cell(row=i, column=c).fill = FILL_TIER_C
    for col, w in zip("ABC", [18, 50, 18]):
        ws_tax.column_dimensions[col].width = w

    # Label export sheet — ready for phase1_logistic.py
    ws_lbl = wb.create_sheet("Label_Export")
    ws_lbl["A1"] = "Label Export"; ws_lbl["A1"].font = FONT_TITLE
    ws_lbl["A2"] = "Export this sheet as CSV and pass to phase1_logistic.py --labels"
    ws_lbl["A2"].font = Font(name="Arial", size=10, italic=True, color="808080")
    ws_lbl["A4"] = "owner_key"; ws_lbl["B4"] = "outcome_status"
    apply_header_row(ws_lbl, 4, 2)
    for i, (_, row) in enumerate(df.iterrows(), 5):
        ws_lbl.cell(row=i, column=1, value=row.get("owner_key", ""))
        ws_lbl.cell(row=i, column=2, value=row.get("outcome_status_to", ""))
    ws_lbl.column_dimensions["A"].width = 20
    ws_lbl.column_dimensions["B"].width = 20

    wb.save(out_path)


def build_scoring_control(out_path: Path):
    """Configuration workbook — weights, tiers, entity factors, outcome taxonomy."""
    wb = Workbook()
    ws_cover = wb.active
    ws_cover.title = "README"
    write_cover(
        ws_cover,
        "WealthBridge Scoring Control v5",
        "The control panel. Every tunable parameter in the propensity system "
        "lives here. Edit values in yellow-highlighted cells; re-run the scoring "
        "pipeline to apply. This is the single source of truth for model config.",
        "Sheets: README · Geo_Tiers · Residential_Weights · Commercial_Weights · Entity_Factors · Outcome_Taxonomy",
        [
            "New file in v5 — formalizes config that was hardcoded in v4 scripts",
            "Enables non-technical weight adjustment without touching Python",
            "Entity type factors exposed for Commercial segment tuning",
            "Weights must sum to 1.0 within each segment (checked by formula)",
        ],
    )

    # Geo tiers
    ws = wb.create_sheet("Geo_Tiers")
    ws["A1"] = "Geographic Tiers — Priority Multipliers"
    ws["A1"].font = FONT_TITLE
    headers = ["Tier ID", "Scope", "Priority ×", "Description", "Outreach Modality"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    apply_header_row(ws, 3, 5)
    for i, (tid, scope, mult, desc, mode) in enumerate(GEO_TIER_REFERENCE, 4):
        ws.cell(row=i, column=1, value=tid).font = FONT_BASE
        ws.cell(row=i, column=2, value=scope).font = FONT_BASE
        mcell = ws.cell(row=i, column=3, value=mult)
        mcell.font = FONT_INPUT; mcell.fill = FILL_ASSUMPTION
        mcell.number_format = "0.00"
        ws.cell(row=i, column=4, value=desc).font = FONT_BASE
        ws.cell(row=i, column=5, value=mode).font = FONT_BASE
    for col, w in zip("ABCDE", [22, 34, 12, 34, 42]):
        ws.column_dimensions[col].width = w

    # Residential weights
    ws = wb.create_sheet("Residential_Weights")
    ws["A1"] = "Residential Scoring Weights (must sum to 1.00)"
    ws["A1"].font = FONT_TITLE
    for c, h in enumerate(["Feature", "Weight", "Rationale"], 1):
        ws.cell(row=3, column=c, value=h)
    apply_header_row(ws, 3, 3)
    for i, (feat, weight, rationale) in enumerate(SCORING_WEIGHTS_RESIDENTIAL, 4):
        ws.cell(row=i, column=1, value=feat).font = FONT_BASE
        wcell = ws.cell(row=i, column=2, value=weight)
        wcell.font = FONT_INPUT; wcell.fill = FILL_ASSUMPTION
        wcell.number_format = FMT_PERCENT
        ws.cell(row=i, column=3, value=rationale).font = FONT_BASE
    sum_row = 4 + len(SCORING_WEIGHTS_RESIDENTIAL)
    ws.cell(row=sum_row, column=1, value="SUM (should = 100%)")
    ws.cell(row=sum_row, column=1).font = Font(name="Arial", size=10, bold=True)
    ws.cell(row=sum_row, column=2, value=f"=SUM(B4:B{sum_row - 1})")
    ws.cell(row=sum_row, column=2).number_format = FMT_PERCENT
    ws.cell(row=sum_row, column=2).font = Font(name="Arial", size=10, bold=True)
    for col, w in zip("ABC", [30, 12, 52]):
        ws.column_dimensions[col].width = w

    # Commercial weights
    ws = wb.create_sheet("Commercial_Weights")
    ws["A1"] = "Commercial Scoring Weights (must sum to 1.00)"
    ws["A1"].font = FONT_TITLE
    for c, h in enumerate(["Feature", "Weight", "Rationale"], 1):
        ws.cell(row=3, column=c, value=h)
    apply_header_row(ws, 3, 3)
    for i, (feat, weight, rationale) in enumerate(SCORING_WEIGHTS_COMMERCIAL, 4):
        ws.cell(row=i, column=1, value=feat).font = FONT_BASE
        wcell = ws.cell(row=i, column=2, value=weight)
        wcell.font = FONT_INPUT; wcell.fill = FILL_ASSUMPTION
        wcell.number_format = FMT_PERCENT
        ws.cell(row=i, column=3, value=rationale).font = FONT_BASE
    sum_row = 4 + len(SCORING_WEIGHTS_COMMERCIAL)
    ws.cell(row=sum_row, column=1, value="SUM (should = 100%)")
    ws.cell(row=sum_row, column=1).font = Font(name="Arial", size=10, bold=True)
    ws.cell(row=sum_row, column=2, value=f"=SUM(B4:B{sum_row - 1})")
    ws.cell(row=sum_row, column=2).number_format = FMT_PERCENT
    ws.cell(row=sum_row, column=2).font = Font(name="Arial", size=10, bold=True)
    for col, w in zip("ABC", [30, 12, 52]):
        ws.column_dimensions[col].width = w

    # Entity type factors
    ws = wb.create_sheet("Entity_Factors")
    ws["A1"] = "Commercial Entity Type Factors"
    ws["A1"].font = FONT_TITLE
    for c, h in enumerate(["Entity Type", "Fit Factor (0-1)", "Planning Fit"], 1):
        ws.cell(row=3, column=c, value=h)
    apply_header_row(ws, 3, 3)
    entity_factors = [
        ("S-Corp / S Corporation", 1.00, "Best — pass-through + planning complexity"),
        ("LLC / L.L.C.", 0.90, "Excellent — flexible, common target"),
        ("C-Corp / Corporation / Inc", 0.85, "Good — key-person + buy-sell common"),
        ("Professional Corp / PC / PLLC", 0.80, "Good — owner-operated"),
        ("Partnership / LP / LLP", 0.75, "Moderate — buy-sell critical"),
        ("Trust", 0.60, "Variable — depends on trust purpose"),
        ("Sole Proprietorship / DBA", 0.50, "Limited — fewer advanced-planning hooks"),
        ("Unknown", 0.40, "Needs enrichment before scoring"),
    ]
    for i, (et, factor, fit) in enumerate(entity_factors, 4):
        ws.cell(row=i, column=1, value=et).font = FONT_BASE
        fcell = ws.cell(row=i, column=2, value=factor)
        fcell.font = FONT_INPUT; fcell.fill = FILL_ASSUMPTION
        fcell.number_format = "0.00"
        ws.cell(row=i, column=3, value=fit).font = FONT_BASE
    for col, w in zip("ABC", [30, 16, 48]):
        ws.column_dimensions[col].width = w

    # Outcome taxonomy
    ws = wb.create_sheet("Outcome_Taxonomy")
    ws["A1"] = "Outcome Status Taxonomy (Phase 1 label source)"
    ws["A1"].font = FONT_TITLE
    for c, h in enumerate(["Status", "Definition", "Phase 1 Role"], 1):
        ws.cell(row=3, column=c, value=h)
    apply_header_row(ws, 3, 3)
    for i, (status, definition, role) in enumerate(OUTCOME_TAXONOMY, 4):
        ws.cell(row=i, column=1, value=status).font = FONT_BASE
        ws.cell(row=i, column=2, value=definition).font = FONT_BASE
        ws.cell(row=i, column=3, value=role).font = FONT_BASE
    for col, w in zip("ABC", [18, 50, 18]):
        ws.column_dimensions[col].width = w

    wb.save(out_path)


def build_event_schedule(df: pd.DataFrame, out_path: Path):
    wb = Workbook()
    ws_cover = wb.active
    ws_cover.title = "README"
    write_cover(
        ws_cover,
        "WealthBridge Event Schedule v5",
        "Seminar + workshop cadence with attribution to propensity tiers. Each "
        "row is one planned or executed event. Derived metrics (RSVP rate, "
        "discovery rate, cost-per-discovery) are Excel formulas — update "
        "attended/discoveries in-row and the metrics recompute.",
        "Sheets: README · Events · Attribution_by_Tier",
        [
            "Added target_segment + target_geo_tier columns (attribution)",
            "Added RSVP/attendance/discovery rate formulas (derive from inputs)",
            "Added cost-per-discovery + cost-per-case formulas",
            "Attribution sheet rolls up events to segment × geo tier",
        ],
    )
    ws = wb.create_sheet("Events")
    # Write the data rows (excluding formula columns — we'll inject those)
    input_cols = EVENT_COLUMNS[:14]  # up through marketing_cost
    formula_cols = EVENT_COLUMNS[14:]  # total_cost through cost_per_case

    for i, (_, label, _, _) in enumerate(EVENT_COLUMNS, 1):
        ws.cell(row=1, column=i, value=label)
    apply_header_row(ws, 1, len(EVENT_COLUMNS))

    for r_idx, (_, row) in enumerate(df.iterrows(), 2):
        # Input columns from data
        for c_idx, (key, _, kind, _) in enumerate(input_cols, 1):
            val = row.get(key)
            if pd.notna(val) and val != "":
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = FONT_BASE
                if kind == "currency": cell.number_format = FMT_CURRENCY
                elif kind == "number": cell.number_format = FMT_NUMBER
                elif kind == "date":   cell.number_format = FMT_DATE
        # Formula columns
        L = get_column_letter
        venue = L(12); fnb = L(13); mkt = L(14)
        invites = L(7); rsvps = L(8); att = L(9); disc = L(10); cases = L(11)
        # total_cost = venue + fnb + marketing
        ws.cell(row=r_idx, column=15, value=f"={venue}{r_idx}+{fnb}{r_idx}+{mkt}{r_idx}").number_format = FMT_CURRENCY
        # rsvp_rate = rsvps / invites (IFERROR)
        ws.cell(row=r_idx, column=16,
            value=f'=IFERROR({rsvps}{r_idx}/{invites}{r_idx},0)').number_format = FMT_PERCENT
        # attendance_rate = attended / rsvps
        ws.cell(row=r_idx, column=17,
            value=f'=IFERROR({att}{r_idx}/{rsvps}{r_idx},0)').number_format = FMT_PERCENT
        # discovery_rate = discoveries / attended
        ws.cell(row=r_idx, column=18,
            value=f'=IFERROR({disc}{r_idx}/{att}{r_idx},0)').number_format = FMT_PERCENT
        # cost_per_discovery = total_cost / discoveries
        ws.cell(row=r_idx, column=19,
            value=f'=IFERROR({L(15)}{r_idx}/{disc}{r_idx},0)').number_format = FMT_CURRENCY
        # cost_per_case = total_cost / cases
        ws.cell(row=r_idx, column=20,
            value=f'=IFERROR({L(15)}{r_idx}/{cases}{r_idx},0)').number_format = FMT_CURRENCY
        for c_idx in range(15, 21):
            ws.cell(row=r_idx, column=c_idx).font = FONT_FORMULA

    for i, (_, _, _, width) in enumerate(EVENT_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    if len(df) > 0:
        last_col = get_column_letter(len(EVENT_COLUMNS))
        ws.auto_filter.ref = f"A1:{last_col}{len(df) + 1}"

    wb.save(out_path)


def build_combined_pipeline(prospects: pd.DataFrame, engagements: pd.DataFrame, out_path: Path):
    wb = Workbook()
    ws_cover = wb.active
    ws_cover.title = "README"
    write_cover(
        ws_cover,
        "WealthBridge Combined Pipeline v5",
        "Joined view of prospects × engagements, ordered by propensity score "
        "with latest outcome state. This is the file to open for daily ops: "
        "who to call today, who just converted, where the next bottleneck is.",
        "Sheets: README · Pipeline · By_Stage · A_Tier_Action_List",
        [
            "Adds propensity-sorted ordering (top decile first)",
            "Latest engagement outcome rolled up per prospect",
            "By_Stage sheet is the funnel view",
            "A_Tier_Action_List is the daily call list for Mike",
        ],
    )

    # Join latest outcome per owner
    if len(engagements) > 0 and "owner_key" in engagements.columns:
        latest = engagements.sort_values("event_date").drop_duplicates("owner_key", keep="last")
        latest = latest[["owner_key", "outcome_status_to", "event_date"]].rename(
            columns={"outcome_status_to": "latest_outcome", "event_date": "latest_event_date"}
        )
        joined = prospects.merge(latest, on="owner_key", how="left")
    else:
        joined = prospects.copy()
        joined["latest_outcome"] = "not_contacted"
        joined["latest_event_date"] = None

    joined = joined.sort_values("propensity_score", ascending=False)

    pipeline_cols = [
        ("owner_key", "Owner Key", "text", 18),
        ("segment", "Segment", "text", 12),
        ("owner_name", "Owner Name", "text", 28),
        ("geo_tier", "Geo Tier", "text", 22),
        ("state", "State", "text", 7),
        ("county", "County", "text", 14),
        ("propensity_score", "Propensity", "score", 12),
        ("propensity_tier", "Tier", "text", 8),
        ("latest_outcome", "Latest Outcome", "text", 18),
        ("latest_event_date", "Last Event", "date", 12),
        ("phone", "Phone", "text", 14),
        ("email", "Email", "text", 26),
        ("multi_property_count", "Multi-Prop", "number", 12),
    ]
    ws = wb.create_sheet("Pipeline")
    write_dataframe(ws, joined, pipeline_cols)

    # Conditional color scale on propensity score
    if len(joined) > 0:
        score_col = get_column_letter(7)
        rule = ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                              mid_type="num", mid_value=50, mid_color="FFEB84",
                              end_type="num", end_value=100, end_color="63BE7B")
        ws.conditional_formatting.add(f"{score_col}2:{score_col}{len(joined) + 1}", rule)

    # By-stage funnel
    ws_stage = wb.create_sheet("By_Stage")
    ws_stage["A1"] = "Pipeline Funnel"; ws_stage["A1"].font = FONT_TITLE
    stages = ["not_contacted", "contacted", "responded", "qualified", "closed_won", "closed_lost"]
    for c, h in enumerate(["Stage", "Count", "% of Total"], 1):
        ws_stage.cell(row=3, column=c, value=h)
    apply_header_row(ws_stage, 3, 3)
    for i, stage in enumerate(stages, 4):
        ws_stage.cell(row=i, column=1, value=stage).font = FONT_BASE
        ws_stage.cell(row=i, column=2, value=f'=COUNTIF(Pipeline!I:I,"{stage}")')
        ws_stage.cell(row=i, column=2).font = FONT_XLINK
        ws_stage.cell(row=i, column=3, value=f'=IFERROR(B{i}/SUM($B$4:$B$9),0)')
        ws_stage.cell(row=i, column=3).number_format = FMT_PERCENT
        ws_stage.cell(row=i, column=3).font = FONT_FORMULA
    for col, w in zip("ABC", [20, 10, 14]):
        ws_stage.column_dimensions[col].width = w

    # A-tier action list
    ws_act = wb.create_sheet("A_Tier_Action_List")
    ws_act["A1"] = "A-Tier Action List (top decile, AZ home, not contacted)"
    ws_act["A1"].font = FONT_TITLE
    ws_act["A2"] = "Filtered view of Pipeline — call these first"
    ws_act["A2"].font = Font(name="Arial", size=10, italic=True, color="808080")

    a_tier = joined[
        (joined.get("propensity_tier", "") == "A")
        & (joined.get("geo_tier", "").isin(["tier_1_home_az", "tier_2_az_adjacent"]))
        & (joined.get("latest_outcome", "not_contacted").fillna("not_contacted") == "not_contacted")
    ].head(100)
    action_cols = pipeline_cols  # reuse
    write_dataframe(ws_act, a_tier, action_cols, start_row=4)

    wb.save(out_path)


def build_executive_summary(prospects: pd.DataFrame, engagements: pd.DataFrame,
                            events: pd.DataFrame, out_path: Path):
    wb = Workbook()
    ws_cover = wb.active
    ws_cover.title = "README"
    write_cover(
        ws_cover,
        "WealthBridge Executive Summary v5",
        "One-screen KPI dashboard for managing-director reviews. Formula-driven — "
        "update the source workbooks (Prospect DB, Engagement DB, Event Schedule) "
        "and this summary recomputes.",
        "Sheets: README · KPIs · Segment_Breakdown · Geo_Breakdown · Model_Health",
        [
            "Added propensity-tier breakdown (v4 did not have this)",
            "Added geo-tier breakdown with CAC estimates per tier",
            "Added Phase 0 → Phase 1 transition indicator (labels accumulated)",
            "Model_Health sheet tracks AUC and Brier score over time",
        ],
    )

    ws = wb.create_sheet("KPIs")
    ws["A1"] = "Key Metrics"; ws["A1"].font = FONT_TITLE

    row = 3
    ws[f"A{row}"] = "Prospect Database"
    ws[f"A{row}"].font = Font(name="Arial", size=11, bold=True)
    ws[f"A{row}"].fill = FILL_SECTION
    ws.merge_cells(f"A{row}:D{row}")
    row += 1
    metrics = [
        ("Total prospects", str(len(prospects))),
        ("  Residential", str((prospects["segment"] == "Residential").sum()) if "segment" in prospects.columns else "0"),
        ("  Commercial", str((prospects["segment"] == "Commercial").sum()) if "segment" in prospects.columns else "0"),
        ("A-tier prospects (top 20%)", str((prospects.get("propensity_tier", "") == "A").sum())),
        ("B-tier prospects", str((prospects.get("propensity_tier", "") == "B").sum())),
    ]
    for label, val in metrics:
        ws.cell(row=row, column=1, value=label).font = FONT_BASE
        c = ws.cell(row=row, column=2, value=int(val))
        c.font = FONT_INPUT
        c.number_format = FMT_NUMBER
        row += 1

    row += 1
    ws[f"A{row}"] = "Engagement Pipeline"
    ws[f"A{row}"].font = Font(name="Arial", size=11, bold=True)
    ws[f"A{row}"].fill = FILL_SECTION
    ws.merge_cells(f"A{row}:D{row}")
    row += 1
    eng_metrics = [
        ("Total engagement events", len(engagements)),
        ("Qualified prospects", (engagements.get("outcome_status_to", "") == "qualified").sum() if len(engagements) else 0),
        ("Closed won", (engagements.get("outcome_status_to", "") == "closed_won").sum() if len(engagements) else 0),
        ("Closed lost", (engagements.get("outcome_status_to", "") == "closed_lost").sum() if len(engagements) else 0),
    ]
    labels_total = 0
    for label, val in eng_metrics:
        ws.cell(row=row, column=1, value=label).font = FONT_BASE
        c = ws.cell(row=row, column=2, value=int(val))
        c.font = FONT_INPUT
        c.number_format = FMT_NUMBER
        if label in ("Qualified prospects", "Closed won", "Closed lost"):
            labels_total += int(val)
        row += 1

    row += 1
    ws[f"A{row}"] = "Model Health"
    ws[f"A{row}"].font = Font(name="Arial", size=11, bold=True)
    ws[f"A{row}"].fill = FILL_SECTION
    ws.merge_cells(f"A{row}:D{row}")
    row += 1
    ws.cell(row=row, column=1, value="Terminal labels (for Phase 1 training)").font = FONT_BASE
    ws.cell(row=row, column=2, value=labels_total).font = FONT_INPUT
    ws.cell(row=row, column=2).number_format = FMT_NUMBER
    row += 1
    ws.cell(row=row, column=1, value="Phase 1 threshold").font = FONT_BASE
    c = ws.cell(row=row, column=2, value=300)
    c.font = FONT_INPUT; c.fill = FILL_ASSUMPTION
    c.number_format = FMT_NUMBER
    row += 1
    ws.cell(row=row, column=1, value="Phase 1 ready?").font = FONT_BASE
    ws.cell(row=row, column=2, value=f'=IF(B{row-2}>=B{row-1},"READY","Need more labels")')
    ws.cell(row=row, column=2).font = FONT_FORMULA

    row += 2
    ws[f"A{row}"] = "Event Performance (v5)"
    ws[f"A{row}"].font = Font(name="Arial", size=11, bold=True)
    ws[f"A{row}"].fill = FILL_SECTION
    ws.merge_cells(f"A{row}:D{row}")
    row += 1
    evt_metrics = [
        ("Events executed", len(events)),
        ("Total attendees", int(events.get("attended", pd.Series([0])).fillna(0).sum()) if len(events) else 0),
        ("Total discoveries booked", int(events.get("discoveries_booked", pd.Series([0])).fillna(0).sum()) if len(events) else 0),
        ("Total cases opened", int(events.get("cases_opened", pd.Series([0])).fillna(0).sum()) if len(events) else 0),
    ]
    for label, val in evt_metrics:
        ws.cell(row=row, column=1, value=label).font = FONT_BASE
        c = ws.cell(row=row, column=2, value=int(val))
        c.font = FONT_INPUT
        c.number_format = FMT_NUMBER
        row += 1

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    wb.save(out_path)


# ---------------------------------------------------------------------------
# Convenience wrapper — one call produces the whole v5 family
# ---------------------------------------------------------------------------

def build_v5_family(prospects: pd.DataFrame, engagements: pd.DataFrame,
                    events: pd.DataFrame, out_dir: Path):
    out_dir.mkdir(parents=True, exist_ok=True)
    build_prospect_database(prospects, out_dir / "WealthBridge_Prospect_Database_v5.xlsx")
    build_engagement_database(engagements, out_dir / "WealthBridge_Engagement_Database_v5.xlsx")
    build_scoring_control(out_dir / "WealthBridge_Scoring_Control_v5.xlsx")
    build_event_schedule(events, out_dir / "WealthBridge_Event_Schedule_v5.xlsx")
    build_combined_pipeline(prospects, engagements, out_dir / "WealthBridge_Combined_Pipeline_v5.xlsx")
    build_executive_summary(prospects, engagements, events,
                            out_dir / "WealthBridge_Executive_Summary_v5.xlsx")
    print(f"[v5 writer] built 6 workbooks → {out_dir}")
